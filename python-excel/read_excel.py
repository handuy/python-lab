import openpyxl

def readExcel(excelFile, result):
    wb = openpyxl.load_workbook(excelFile)

    activeSheet = wb.active

    dict_item = {
        "Name": "",
        # "FutureState": ""
    }
    # iterate through excel and display data
    for i in range(2, activeSheet.max_row+1):
        for j in range(1, activeSheet.max_column+1):
            col_name = activeSheet.cell(row=1, column=j).value
            
            if col_name == "Name" :
                col_value = activeSheet.cell(row=i, column=j).value
                dict_item["Name"] = col_value
                
            # elif col_name == "FutureState":
            #     col_value = activeSheet.cell(row=i, column=j).value
            #     dict_item["FutureState"] = col_value

            else:
                continue

        result.append(dict_item)
        dict_item = {
            "Name": "",
            # "FutureState": ""
        }
