import openpyxl

def write_excel(list_of_dict, fileName):
    book = openpyxl.Workbook()
    sheet = book.active

    sheet.cell(row = 1, column = 1).value = "Name"
    sheet.cell(row = 1, column = 2).value = "FutureState"

    newRow = 2
    #iterate the each key-value pair of dictionary & insert into sheet

    for dict_item in list_of_dict:
        sheet.cell(row = newRow, column = 1).value = dict_item["Name"]
        sheet.cell(row = newRow, column = 2).value = dict_item["FutureState"]
        newRow = newRow + 1

    book.save(fileName)